#!/usr/bin/env python3
"""
merge_translated_to_xlsx.py
===========================
Merge translated JSONL back into a copy of the original xlsx, appending a text_en column.

Usage:
    python merge_translated_to_xlsx.py \
        --original "datasets/hadith-data/sheikahmad/Hadith-Pure-Canon-Authentica.xlsx" \
        --translated datasets/hadith-data/pure_canon_en.jsonl \
        --output datasets/hadith-data/pure_canon_translated.xlsx \
        --match-col 0 \
        --id-field row_no
"""

import argparse
import io
import json
import sys

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")


def main():
    parser = argparse.ArgumentParser(description="Merge translated JSONL → xlsx")
    parser.add_argument("--original", required=True, help="Original xlsx")
    parser.add_argument("--translated", required=True, help="Translated JSONL")
    parser.add_argument("--output", required=True, help="Output xlsx path")
    parser.add_argument("--sheet", default="Sheet1", help="Sheet name")
    parser.add_argument(
        "--match-col",
        type=int,
        required=True,
        help="0-indexed column in xlsx to match against --id-field",
    )
    parser.add_argument(
        "--id-field",
        default="id",
        help="JSONL field to match against --match-col",
    )
    parser.add_argument(
        "--text-field",
        default="text_en",
        help="JSONL field containing translation",
    )
    args = parser.parse_args()

    # Load translations into lookup
    print(f"Loading translations from {args.translated}...")
    translations = {}
    with open(args.translated, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            obj = json.loads(line)
            key = str(obj.get(args.id_field, ""))
            text = obj.get(args.text_field, "")
            if key and text:
                translations[key] = text
    print(f"  {len(translations)} translations loaded.")

    # Open original xlsx (NOT read_only — we need to modify)
    print(f"Opening {args.original}...")
    wb = openpyxl.load_workbook(args.original)
    ws = wb[args.sheet]

    # Add header for text_en in the next empty column
    max_col = ws.max_column
    en_col = max_col + 1
    ws.cell(row=1, column=en_col, value="text_en")

    matched = 0
    for row_idx in range(2, ws.max_row + 1):
        key = str(ws.cell(row=row_idx, column=args.match_col + 1).value or "").strip()
        if key in translations:
            ws.cell(row=row_idx, column=en_col, value=translations[key])
            matched += 1

        if row_idx % 10000 == 0:
            print(f"  Processed {row_idx} rows...")

    print(f"Matched {matched}/{ws.max_row - 1} rows.")
    print(f"Saving to {args.output}...")
    wb.save(args.output)
    wb.close()
    print("Done.")


if __name__ == "__main__":
    main()
